"""Unit tests for vision OCR helpers and attachment markers."""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock

import pytest

from agent.state import looks_like_file_driven_task
from agent.vision_ocr import (
    DOC_CALIBRATION,
    DOC_GENERIC,
    DOC_LEVELLING,
    DOC_SPREADSHEET,
    DOC_TRAVERSE,
    DOC_UI,
    VisionOcrResult,
    _system_prompt_for_mode,
    _user_prompt_for_ocr,
    apply_survey_context_repairs,
    assess_image_quality,
    build_ocr_review,
    classify_ocr_document,
    export_ocr_extraction_to_excel,
    extract_image_paths_from_query,
    fields_needing_reread,
    format_ocr_review_for_user,
    hz_pair_ok,
    is_ocr_export_request,
    is_ocr_followup_request,
    is_ocr_only_request,
    load_handwriting_style,
    load_last_ocr_extraction,
    looks_like_survey_sheet,
    normalize_extracted_document,
    parse_dms_to_deg,
    parse_survey_number,
    resolve_ocr_export_path,
    run_vision_ocr,
    save_handwriting_style,
    save_last_ocr_extraction,
    select_vision_ocr_mode,
    should_fastpath_image_survey_replot,
    validate_levelling_or_calibration,
    validate_ocr_extraction,
    validate_traverse_rows,
    va_pair_ok,
)
from survyai.attachments import (
    DEFAULT_ATTACHMENTS_ONLY_PROMPT,
    format_attachments_block,
    format_user_transcript,
    parse_attachments_block,
)
from survyai.provider_models import (
    provider_supports_vision,
    vision_unsupported_user_message,
)


def test_format_and_parse_attachments_block_roundtrip(tmp_path: Path) -> None:
    img = tmp_path / "scan.png"
    img.write_bytes(b"fake")
    block = format_attachments_block([str(img)], "OCR this image")
    paths, text = parse_attachments_block(block)
    assert paths == [str(img)]
    assert text == "OCR this image"
    assert "[SurvyAI attachments]" in block


def test_attachments_only_uses_default_prompt() -> None:
    block = format_attachments_block([r"C:\tmp\a.png"], "")
    assert DEFAULT_ATTACHMENTS_ONLY_PROMPT in block
    _, text = parse_attachments_block(block)
    assert text == DEFAULT_ATTACHMENTS_ONLY_PROMPT


def test_format_user_transcript_shows_filenames() -> None:
    out = format_user_transcript("Hello", [r"C:\docs\plan.png"])
    assert "Hello" in out
    assert "plan.png" in out
    assert "base64" not in out.lower()


def test_looks_like_file_driven_task_for_images() -> None:
    assert looks_like_file_driven_task("Please OCR C:\\data\\scan.png")
    assert looks_like_file_driven_task("see attached photo.jpg")
    assert looks_like_file_driven_task("[SurvyAI attachments]\nx.png\n[/SurvyAI attachments]")
    assert not looks_like_file_driven_task("What is a bearing?")


def test_extract_image_paths_from_marker_and_typed(tmp_path: Path) -> None:
    img = tmp_path / "photo.jpg"
    img.write_bytes(b"x")
    doc = tmp_path / "notes.pdf"
    doc.write_bytes(b"%PDF")

    marker_q = format_attachments_block([str(img), str(doc)], "read it")
    found = extract_image_paths_from_query(marker_q)
    assert found == [str(img.resolve())]

    typed = f"Please OCR {img}"
    found2 = extract_image_paths_from_query(typed)
    assert found2 == [str(img.resolve())]


def test_select_vision_ocr_mode() -> None:
    assert select_vision_ocr_mode("OCR this and extract the text") == "plain_text"
    assert select_vision_ocr_mode("extract bearings and pillars from this scan") == "geospatial"
    assert select_vision_ocr_mode("extract all components as structured JSON") == "structured"
    assert select_vision_ocr_mode("process the attached image") == "structured"
    assert select_vision_ocr_mode("OCR this image") == "structured"
    assert select_vision_ocr_mode("OCR this instrument calibration sheet") == "structured"


def test_is_ocr_only_request() -> None:
    assert is_ocr_only_request("OCR this image")
    assert is_ocr_only_request("what does this say")
    assert is_ocr_only_request("Scan through the image attached and give all key details")
    assert is_ocr_only_request("Process the attached file(s).")
    assert is_ocr_only_request("Scan the image and save to excel file e.xlsx")
    assert not is_ocr_only_request("OCR this then replot to DWG")
    assert not is_ocr_only_request("Save the updates to excel file e.xlsx")
    assert not is_ocr_only_request("Summarize those values")


def test_should_fastpath_image_survey_replot(tmp_path: Path) -> None:
    img = tmp_path / "cadastral_plan.png"
    img.write_bytes(b"x")
    q = (
        f"[SurvyAI attachments]\n{img}\n[/SurvyAI attachments]\n\n"
        "Replot this cadastral survey plan to C:\\out\\plan.dwg"
    )
    assert should_fastpath_image_survey_replot(q)
    assert not should_fastpath_image_survey_replot(
        f"[SurvyAI attachments]\n{img}\n[/SurvyAI attachments]\n\nJust OCR the text"
    )


def test_provider_vision_capability_matrix() -> None:
    assert provider_supports_vision("openai")
    assert provider_supports_vision("claude")
    assert provider_supports_vision("gemini")
    assert not provider_supports_vision("deepseek")
    assert not provider_supports_vision("ollama")
    msg = vision_unsupported_user_message("deepseek")
    assert "deepseek" in msg.lower()
    assert "OpenAI" in msg or "openai" in msg.lower()


def test_run_vision_ocr_no_images() -> None:
    result = run_vision_ocr(
        [],
        user_text="OCR",
        llm=MagicMock(),
        run_with_timeout=lambda t, fn: (None, "unused", False),
    )
    assert result.success is False
    assert "No readable image" in (result.error or "")


def test_run_vision_ocr_plain_text_success(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    img = tmp_path / "a.png"
    img.write_bytes(b"not-a-real-png")

    monkeypatch.setattr(
        "agent.vision_ocr.image_file_to_base64_png",
        lambda path, max_edge=2048, enhance=False: "YmFzZTY0",
    )

    class _Msg:
        content = "Hello from OCR"

    def _timeout(t, fn):
        return _Msg(), None, False

    result = run_vision_ocr(
        [str(img)],
        user_text="OCR this image",
        llm=MagicMock(),
        run_with_timeout=_timeout,
        workspace=tmp_path,
        mode="plain_text",
        model_name="test-model",
    )
    assert result.success is True
    assert result.text == "Hello from OCR"
    assert result.model_name == "test-model"


def test_run_vision_ocr_json_parse_fallback(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    img = tmp_path / "b.png"
    img.write_bytes(b"x")
    monkeypatch.setattr(
        "agent.vision_ocr.image_file_to_base64_png",
        lambda path, max_edge=2048, enhance=False: "YmFzZTY0",
    )

    class _Msg:
        content = "not json but still useful text"

    def _timeout(t, fn):
        return _Msg(), None, False

    result = run_vision_ocr(
        [str(img)],
        user_text="extract structured components",
        llm=MagicMock(),
        run_with_timeout=_timeout,
        workspace=tmp_path,
        mode="structured",
    )
    assert result.success is True
    assert "useful text" in result.text
    assert "JSON parse failed" in (result.notes or "")


def test_run_vision_ocr_structured_json(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    img = tmp_path / "c.png"
    img.write_bytes(b"x")
    monkeypatch.setattr(
        "agent.vision_ocr.image_file_to_base64_png",
        lambda path, max_edge=2048, enhance=False: "YmFzZTY0",
    )
    payload = {
        "plain_text": "Pillar PB1",
        "labels": ["PB1"],
        "coordinates": [{"label": "PB1", "easting": 1.0, "northing": 2.0}],
    }

    class _Msg:
        content = "```json\n" + json.dumps(payload) + "\n```"

    def _timeout(t, fn):
        return _Msg(), None, False

    result = run_vision_ocr(
        [str(img)],
        user_text="extract geospatial components bearings pillars",
        llm=MagicMock(),
        run_with_timeout=_timeout,
        mode="geospatial",
        workspace=tmp_path,
    )
    assert result.success is True
    assert result.structured.get("plain_text") == "Pillar PB1"
    assert result.structured.get("_validation") is not None
    user = result.format_for_user()
    assert "Pillar PB1" in user
    assert "```json" not in user
    assert "PASS" not in user
    assert "FAIL" not in user
    assert "style_card" not in user
    assert "VISION OCR RESULT" in result.format_for_agent_context()


def test_vision_ocr_result_format_failure() -> None:
    r = VisionOcrResult(success=False, error="boom")
    assert "boom" in r.format_for_user()


def test_handwriting_style_lock_in_prompt() -> None:
    prompt = _system_prompt_for_mode("structured", document_type="calibration_sheet")
    assert "style_card" in prompt
    assert "1 vs 7" in prompt or "7" in prompt
    assert "9" in prompt


def test_parse_survey_number_variants() -> None:
    assert parse_survey_number("1.245") == 1.245
    assert parse_survey_number("1.245m") == 1.245
    assert parse_survey_number({"raw": "1,830", "value": 1.83}) == 1.83
    assert parse_survey_number("-") is None


def test_classify_calibration_from_prompt_and_filename() -> None:
    assert classify_ocr_document("scan this instrument calibration sheet") == DOC_CALIBRATION
    assert classify_ocr_document("", [r"C:\docs\two_peg_test.png"]) == DOC_CALIBRATION
    assert looks_like_survey_sheet("backsight and foresight booking")
    assert classify_ocr_document("level book staff readings") == DOC_LEVELLING
    assert classify_ocr_document("give all key details", [r"C:\docs\Calibration.png"]) == DOC_GENERIC
    assert classify_ocr_document(
        "scan the image",
        [r"C:\docs\Calibration.png"],
        {"title": "TRAVERSE ANGLE AND DISTANCE FIELD SHEET", "document_type": "calibration_sheet"},
    ) == DOC_TRAVERSE


def test_hi_equals_rl_plus_bs_catches_digit_error() -> None:
    rows = [
        {
            "station": "A",
            "reduced_level": {"raw": "100.000", "value": 100.0, "confidence": 0.99},
            "backsight": {"raw": "1.245", "value": 1.245, "confidence": 0.98},
            "height_of_collimation": {"raw": "101.245", "value": 101.245, "confidence": 0.99},
        },
        {
            "station": "B",
            "foresight": {"raw": "1.330", "value": 1.330, "confidence": 0.91},
            "reduced_level": {"raw": "99.415", "value": 99.415, "confidence": 0.90},
        },
    ]
    names = {c.name: c.passed for c in validate_levelling_or_calibration(rows)}
    assert names.get("row[0].HI=RL+BS") is True
    assert names.get("row[1].RL=HI-FS") is False


def test_validate_ocr_extraction_flags_review() -> None:
    payload = {
        "document_type": "calibration_sheet",
        "rows": [
            {
                "station": "B",
                "backsight": {"raw": "1.245", "value": 1.245, "confidence": 0.98},
                "intermediate_sight": {"raw": "1.830", "value": 1.330, "confidence": 0.70},
                "foresight": {"raw": "2.115", "value": 2.115, "confidence": 0.97},
                "distance": {"raw": "30.0", "value": 30.0, "confidence": 0.99},
            }
        ],
        "confidence": {"intermediate_sight": 0.70},
    }
    report = validate_ocr_extraction(payload, user_text="calibration sheet")
    assert report.document_type == DOC_CALIBRATION
    assert report.review_required is True
    assert "intermediate_sight" in report.uncertain_fields
    assert "intermediate_sight" in fields_needing_reread(report)


def test_consistent_calibration_row_passes() -> None:
    payload = {
        "document_type": "levelling_book",
        "rows": [
            {"station": "A", "reduced_level": 10.0, "backsight": 1.5, "height_of_collimation": 11.5},
            {"station": "B", "foresight": 2.0, "reduced_level": 9.5},
        ],
        "confidence": {"backsight": 0.99, "foresight": 0.99},
    }
    report = validate_ocr_extraction(payload)
    assert report.ok is True
    assert report.review_required is False


def test_traverse_distance_reconstruct() -> None:
    rows = [{"distance": 5.0, "latitude": 3.0, "departure": 4.0}]
    assert validate_traverse_rows(rows)[0].passed is True
    assert validate_traverse_rows([{"distance": 9.0, "latitude": 3.0, "departure": 4.0}])[0].passed is False


def test_parse_dms_and_face_identities() -> None:
    zero = parse_dms_to_deg("000°00′00″")
    assert zero is not None and abs(zero) < 1e-9
    assert abs((parse_dms_to_deg("179°59′59″") or 0) - 179.999722) < 1e-3
    assert hz_pair_ok(parse_dms_to_deg("000°00′00″"), parse_dms_to_deg("179°59′59″"))
    assert va_pair_ok(parse_dms_to_deg("89°46′23″"), parse_dms_to_deg("270°13′27″"))
    assert not hz_pair_ok(parse_dms_to_deg("000°00′00″"), parse_dms_to_deg("270°00′01″"))


def test_traverse_context_repairs_faint_glyphs() -> None:
    payload = {
        "document_type": "traverse_sheet",
        "title": "TRAVERSE ANGLE AND DISTANCE FIELD SHEET",
        "style_card": {"7": "Angular form with a horizontal cross-stroke", "6": "Closed loop, high stem"},
        "metadata": {
            "instrument": "Geomato MTS-1202RT",
            "serial": "1258133",
            "date": "30/01/2025",
            "surveyed_by": "VICTOR MOKWENYE",
        },
        "rows": [
            {
                "from": "S/AX-03",
                "to": "S/AX-01",
                "hz_fl": "000°00′00″",
                "hz_fr": "179°59′59″",
                "va_fl": "89°46′23″",
                "va_fr": "270°13′27″",
                "slope": 294.651,
                "hor": 294.649,
            },
            {
                "to": "S/AX-02",
                "hz_fl": "182°03′37″",
                "hz_fr": "002°03′35″",
                "va_fl": "90°15′01″",
                "va_fr": "269°44′55″",
                "slope": 343.087,
                "hor": 363.050,
            },
            {
                "from": "S/AX-03",
                "to": "S/AX-01",
                "hz_fl": "000°00′10″",
                "hz_fr": "270°00′01″",
                "va_fl": "89°46′23″",
                "va_fr": "270°13′29″",
                "slope": 294.651,
                "hor": 294.649,
            },
            {
                "to": "S/AX-02",
                "hz_fl": "272°03′39″",
                "hz_fr": "092°03′41″",
                "va_fl": "90°15′05″",
                "va_fr": "269°44′53″",
                "slope": 363.054,
                "hor": 363.050,
            },
            {
                "from": "S/AX-09",
                "to": "S/AX-01",
                "hz_fl": "180°00′01″",
                "hz_fr": "359°59′58″",
                "va_fl": "89°46′24″",
                "va_fr": "270°13′37″",
                "slope": 294.651,
                "hor": 294.649,
            },
            {
                "to": "S/AX-02",
                "hz_fl": "002°03′43″",
                "hz_fr": "182°03′41″",
                "va_fl": "90°15′02″",
                "va_fr": "269°44′52″",
                "slope": 343.087,
                "hor": 363.050,
            },
        ],
    }
    out = apply_survey_context_repairs(payload, document_type=DOC_TRAVERSE)
    meta = out["metadata"]
    assert "1202R+" in str(meta.get("instrument") if not isinstance(meta.get("instrument"), dict) else meta["instrument"].get("value"))
    serial = meta.get("serial")
    serial_v = serial.get("value") if isinstance(serial, dict) else serial
    assert str(serial_v) == "1250033"
    date = meta.get("date")
    date_v = date.get("value") if isinstance(date, dict) else date
    assert str(date_v) == "30/07/2025"
    rows = out["rows"]
    assert all(str(r.get("instrument_station") or r.get("from")).replace("S/AX", "SIAX").startswith("SIAX-03") or "SIAX-03" in str(r) for r in rows)
    assert "SIAX-03" in str(rows[0].get("instrument_station") or rows[0].get("from"))
    assert "SIAX-03" in str(rows[4].get("instrument_station") or rows[4].get("from"))
    fl3 = str(rows[2].get("hz_fl"))
    assert fl3.startswith("090") or "090" in fl3
    slopes = [r.get("slope_distance", r.get("slope")) for r in rows]
    s2 = slopes[1].get("value") if isinstance(slopes[1], dict) else slopes[1]
    s6 = slopes[5].get("value") if isinstance(slopes[5], dict) else slopes[5]
    assert abs(float(s2) - 363.054) < 0.02 or abs(float(s2) - 363.087) < 0.02 or abs(float(s2) - 363.05) < 0.2
    assert abs(float(s6) - 363.054) < 0.02 or abs(float(s6) - 363.087) < 0.02 or abs(float(s6) - 363.05) < 0.2
    user = VisionOcrResult(success=True, structured=out, image_paths=["Calibration.png"]).format_for_user()
    assert "```json" not in user
    assert "PASS" not in user
    assert "style_card" not in user
    assert "1250033" in user
    assert "Geomato MTS-1202R+" in user


def test_handwriting_style_persists(tmp_path: Path) -> None:
    save_handwriting_style({"7": "cross-barred angular 7", "0": "narrow oval"}, workspace=tmp_path, writer="Victor")
    loaded = load_handwriting_style(tmp_path)
    assert "cross" in loaded.get("7", "").lower()
    assert "oval" in loaded.get("0", "").lower()


def test_quality_gate_rejects_tiny_black_image_without_llm(tmp_path: Path) -> None:
    from PIL import Image

    img = tmp_path / "tiny_black.png"
    Image.new("RGB", (40, 40), color=(0, 0, 0)).save(img)
    q = assess_image_quality(str(img))
    assert q["readable"] is False
    assert q["overall"] == "bad"

    llm = MagicMock()

    def _timeout(t, fn):
        raise AssertionError("LLM must not be invoked for unreadable images")

    result = run_vision_ocr(
        [str(img)],
        user_text="OCR this image",
        llm=llm,
        run_with_timeout=_timeout,
        mode="structured",
        workspace=tmp_path,
    )
    assert result.success is False
    assert "retake" in (result.error or "").lower() or "clear" in (result.error or "").lower()
    llm.invoke.assert_not_called()


def test_build_ocr_review_payload_shape() -> None:
    structured = {
        "document_type": "traverse_sheet",
        "title": "TRAVERSE ANGLE AND DISTANCE FIELD SHEET",
        "style_card": {"7": "cross-bar"},
        "metadata": {"serial": {"raw": "1250033", "value": "1250033", "confidence": 0.9}},
        "rows": [
            {
                "instrument_station": {"value": "SIAX-03", "confidence": 0.8, "bbox": [0.1, 0.2, 0.2, 0.25]},
                "reference_station": "SIAX-01",
                "hz_fl": "000°00′00″",
                "hz_fr": "179°59′59″",
                "slope_distance": 294.651,
            }
        ],
    }
    validation = {"uncertain_fields": ["serial"], "checks": []}
    review = build_ocr_review(
        image_paths=[r"C:\docs\Calibration.png"],
        document_type="traverse_sheet",
        model_name="test-model",
        structured=structured,
        validation=validation,
        quality={"overall": "good", "readable": True},
    )
    assert review["image_paths"] == [r"C:\docs\Calibration.png"]
    assert review["document_type"] == "traverse_sheet"
    assert review["model_name"] == "test-model"
    assert "serial" in review["metadata"]
    assert len(review["rows"]) == 1
    assert review["rows"][0]["instrument_station"]["bbox"] == [0.1, 0.2, 0.2, 0.25]
    assert "serial" in review["uncertain"]
    user = VisionOcrResult(
        success=True, structured=structured, image_paths=["Calibration.png"], ocr_review=review
    ).format_for_user()
    assert "PASS" not in user
    assert "```json" not in user
    assert "SIAX-03" in user or "294.651" in user


def test_run_vision_ocr_attaches_ocr_review(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from PIL import Image

    img = tmp_path / "sheet.png"
    # Enough detail that quality gate passes
    im = Image.new("RGB", (400, 500), color=(220, 220, 220))
    for x in range(0, 400, 3):
        for y in range(0, 500, 7):
            im.putpixel((x, y), (40, 40, 40))
    im.save(img)

    monkeypatch.setattr(
        "agent.vision_ocr.image_file_to_base64_png",
        lambda path, max_edge=2048, enhance=False: "YmFzZTY0",
    )
    payload = {
        "document_type": "traverse_sheet",
        "title": "TRAVERSE SHEET",
        "metadata": {"date": "30/07/2025"},
        "rows": [{"from": "SIAX-03", "to": "SIAX-01", "hz_fl": "000°00′00″", "hz_fr": "180°00′00″"}],
    }

    class _Msg:
        content = "```json\n" + json.dumps(payload) + "\n```"

    def _timeout(t, fn):
        return _Msg(), None, False

    result = run_vision_ocr(
        [str(img)],
        user_text="give all key details",
        llm=MagicMock(),
        run_with_timeout=_timeout,
        mode="structured",
        workspace=tmp_path,
    )
    assert result.success is True
    assert result.ocr_review
    assert result.ocr_review.get("rows")
    assert result.quality.get("readable") is True
    assert "PASS" not in result.format_for_user()


def test_normalize_nested_traverse_angles_and_phone() -> None:
    from agent.vision_ocr import format_ocr_review_for_user, normalize_extracted_document

    raw = {
        "document_type": "traverse_sheet",
        "title": "TRAVERSE ANGLE AND DISTANCE FIELD SHEET",
        "metadata": {
            "organization": "West Field Energy Resources Ltd.",
            "phone": "070 31083907",
            "instrument": "Geomato MTS-1202R+",
        },
        "rows": [
            {
                "from": "SIAX-03",
                "to": "SIAX-01",
                "horizontal_angle": {
                    "face_left": {"deg": 0, "min": 0, "sec": 0},
                    "face_right": "179°59′59″",
                },
                "vertical_angle": {"fl": "89°46′23″", "fr": "270°13′27″"},
                "distance": {"slope": 294.651, "hor": 294.649},
            }
        ],
    }
    out = normalize_extracted_document(raw)
    row = out["rows"][0]
    assert "000°" in str(row.get("hz_fl")) or str(row.get("hz_fl")).startswith("0")
    assert "179" in str(row.get("hz_fr"))
    assert "89" in str(row.get("va_fl"))
    assert "270" in str(row.get("va_fr"))
    assert float(row.get("slope_distance") or 0) == 294.651
    user = VisionOcrResult(success=True, structured=out, image_paths=["Calibration.png"]).format_for_user()
    assert "070 31083907" in user
    assert "Face Left" in user or "HA" in user
    assert "294.651" in user

    review = build_ocr_review(
        image_paths=["Calibration.png"],
        document_type="traverse_sheet",
        model_name="t",
        structured=out,
        validation={},
        quality={"overall": "good"},
    )
    review["metadata"]["serial"] = {"value": "1250033", "raw": "1250033"}
    updated = format_ocr_review_for_user(review, note="Updated 1 field(s) from image review.")
    assert "Updated 1 field(s)" in updated
    assert "1250033" in updated
    assert "294.651" in updated
    assert "070 31083907" in updated


def test_repair_all_inst_stn_misread_as_ref_and_343_to_363() -> None:
    """Consistent 2↔3 / 4↔6 misreads must not survive majority vote."""
    payload = {
        "metadata": {
            "instrument": "Geomato MTS-1202R",
            "serial": "125123",
            "surveyed_by": "Victor Mokwenoye",
            "phone": "070…",
            "date": "30/07/2025",
        },
        "rows": [
            {"from": "SIAX-02", "to": "SIAX-04", "hz_fl": "000°00′00″", "hz_fr": "179°59′59″",
             "va_fl": "89°46′23″", "va_fr": "270°13′27″", "slope": 294.651, "hor": 294.649},
            {"from": "SIAX-02", "to": "SIAX-02", "hz_fl": "182°03′37″", "hz_fr": "002°03′35″",
             "va_fl": "90°15′01″", "va_fr": "269°44′55″", "slope": 343.054, "hor": 343.050},
            {"from": "SIAX-02", "to": "SIAX-04", "hz_fl": "090°00′00″", "hz_fr": "270°00′01″",
             "va_fl": "89°46′28″", "va_fr": "270°13′29″", "slope": 294.651, "hor": 294.649},
            {"from": "SIAX-02", "to": "SIAX-02", "hz_fl": "272°03′39″", "hz_fr": "092°03′41″",
             "va_fl": "90°15′05″", "va_fr": "269°44′53″", "slope": 343.054, "hor": 343.050},
            {"from": "SIAX-02", "to": "SIAX-04", "hz_fl": "180°00′01″", "hz_fr": "359°59′58″",
             "va_fl": "89°46′24″", "va_fr": "270°13′37″", "slope": 294.651, "hor": 294.649},
            {"from": "SIAX-02", "to": "SIAX-02", "hz_fl": "002°03′43″", "hz_fr": "182°03′41″",
             "va_fl": "90°15′06″", "va_fr": "269°44′52″", "slope": 343.053, "hor": 343.050},
        ],
    }
    out = apply_survey_context_repairs(payload, document_type=DOC_TRAVERSE)
    meta = out["metadata"]
    assert "1202R+" in str(meta.get("instrument") if not isinstance(meta.get("instrument"), dict) else meta["instrument"].get("value"))
    serial = meta.get("serial")
    serial_v = serial.get("value") if isinstance(serial, dict) else serial
    assert str(serial_v) == "1250033"
    surveyed = meta.get("surveyed_by")
    surveyed_v = surveyed.get("value") if isinstance(surveyed, dict) else surveyed
    assert "Mokwenye" in str(surveyed_v)
    assert "Mokwenoye" not in str(surveyed_v)
    rows = out["rows"]
    for row in rows:
        inst = str(row.get("instrument_station") or row.get("from"))
        assert "SIAX-03" in inst
        assert "SIAX-02" not in inst or "SIAX-03" in inst
    slopes = [float(r.get("slope_distance") or r.get("slope")) for r in rows]
    assert abs(slopes[1] - 363.054) < 0.02
    assert abs(slopes[5] - 363.053) < 0.02


def test_repair_244_to_294_and_hor_align() -> None:
    payload = {
        "rows": [
            {"from": "SIAX-03", "to": "SIAX-04", "slope": 244.651, "hor": 244.649},
            {"from": "SIAX-03", "to": "SIAX-02", "slope": 363.057, "hor": 362.65},
            {"from": "SIAX-03", "to": "SIAX-04", "slope": 244.651, "hor": 244.649},
            {"from": "SIAX-03", "to": "SIAX-02", "slope": 363.057, "hor": 362.65},
        ],
    }
    out = apply_survey_context_repairs(payload, document_type=DOC_TRAVERSE)
    rows = out["rows"]
    s0 = float(rows[0].get("slope_distance") or rows[0].get("slope"))
    h0 = float(rows[0].get("horizontal_distance") or rows[0].get("hor"))
    s1 = float(rows[1].get("slope_distance") or rows[1].get("slope"))
    h1 = float(rows[1].get("horizontal_distance") or rows[1].get("hor"))
    assert abs(s0 - 294.651) < 0.02
    assert abs(h0 - 294.649) < 0.02
    assert abs(s1 - 363.057) < 0.02
    assert abs(h1 - s1) <= 0.5


def test_spreadsheet_screenshot_format_and_excel(tmp_path: Path) -> None:
    from agent.vision_ocr import DOC_SPREADSHEET, classify_ocr_document, normalize_extracted_document

    raw = {
        "document_type": "generic_spreadsheet",
        "title": "DELETE - Excel",
        "style_card": {"0": "narrow oval"},
        "metadata": {
            "organization": None,
            "phone": None,
            "surveyed_by": None,
            "instrument": None,
            "serial": None,
            "sheet": "tmp_geom_rows_batch_001_2025121",
        },
        "rows": [],
        "visible_headers": [
            "ATTR_BE", "ATTR_EN", "ATTR_ON", "ATTR_DI", "ATTR_DA",
            "ATTR_ORIGINAL_", "ATTR_AT", "ATTR_Sh", "PART_INDEX", "VERTEX_IX", "X", "Y", "Z", "M",
        ],
        "visible_records": [
            {
                "excel_row": 2, "ATTR_ON": "ONS", "ATTR_DI": "EAST", "ATTR_DA": "GCS",
                "ATTR_ORIGINAL_": "Nig M Belt", "ATTR_AT": "ABIMBOLA",
                "date_time": "2014-04-14 00:00:00", "ATTR_Sh": 0.1987789,
                "PART_INDEX": 0, "VERTEX_IX": 0, "X": 7.128806, "Y": 4.840182, "Z": 17.4, "M": 5068,
            },
            {
                "excel_row": 3, "ATTR_ON": "ONS", "ATTR_DI": "EAST", "ATTR_DA": "GCS",
                "ATTR_ORIGINAL_": "Nig M Belt", "ATTR_AT": "ABIMBOLA",
                "date_time": "2014-04-14 00:00:00", "ATTR_Sh": 0.1987789,
                "PART_INDEX": 0, "VERTEX_IX": 1, "X": 7.20859, "Y": 4.84034, "Z": 20.9, "M": 5245,
            },
        ],
    }
    out = normalize_extracted_document(raw)
    assert classify_ocr_document("", ["Screenshot (8).png"], out) == DOC_SPREADSHEET
    assert len(out["rows"]) == 2
    assert out["rows"][0]["X"] == 7.128806
    # Null survey metadata stripped
    assert "organization" not in (out.get("metadata") or {}) or out["metadata"].get("organization") not in (None, "")
    assert (out.get("metadata") or {}).get("sheet")

    user = VisionOcrResult(
        success=True, structured=out, image_paths=["Screenshot (8).png"], document_type=DOC_SPREADSHEET
    ).format_for_user()
    assert "style_card" not in user
    assert "Surveyed by" not in user
    assert "Instrument" not in user
    assert "Inst. Stn" not in user
    assert "ATTR_ON" in user
    assert "7.128806" in user
    assert "ONS" in user
    assert not user.strip().startswith("{")

    xlsx = export_ocr_extraction_to_excel(out, tmp_path / "f.xlsx", image_paths=["Screenshot (8).png"])
    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    assert wb.sheetnames[0] == "Extraction"
    obs = wb["Observations"]
    headers = [c.value for c in obs[1]]
    assert "ATTR_ON" in headers
    assert "X" in headers
    assert "Y" in headers
    assert obs.max_row >= 3
    assert obs[2][headers.index("ATTR_ON")].value == "ONS"
    assert float(obs[2][headers.index("X")].value) == pytest.approx(7.128806)


def test_ocr_excel_export_and_last_extraction(tmp_path: Path) -> None:
    assert is_ocr_export_request("Save the updates to excel file e.xlsx")
    assert is_ocr_followup_request("Save the updates to excel file e.xlsx")
    structured = {
        "title": "TRAVERSE ANGLE AND DISTANCE FIELD SHEET",
        "metadata": {
            "organization": "West Field Energy Resources Ltd.",
            "surveyed_by": "Victor Mokwenye",
            "serial": "1250033",
        },
        "rows": [
            {"from": "SIAX-03", "to": "SIAX-04", "hz_fl": "000°00′00″", "hz_fr": "179°59′59″",
             "slope": 294.651, "hor": 294.649},
            {"from": "SIAX-03", "to": "SIAX-02", "hz_fl": "182°03′37″", "hz_fr": "002°03′35″",
             "slope": 363.054, "hor": 363.050},
        ],
    }
    saved = save_last_ocr_extraction(
        structured, workspace=tmp_path, image_paths=["Calibration.png"], document_type=DOC_TRAVERSE
    )
    assert saved is not None and saved.is_file()
    loaded = load_last_ocr_extraction(tmp_path)
    assert loaded is not None
    assert loaded["structured"]["metadata"]["serial"] == "1250033"

    out = resolve_ocr_export_path("Save the updates to excel file e.xlsx", tmp_path)
    assert out.name == "e.xlsx"
    assert out.parent == tmp_path.resolve()
    # File must not need to exist beforehand
    assert not out.exists()
    written = export_ocr_extraction_to_excel(structured, out, image_paths=["Calibration.png"])
    assert written.is_file()
    from openpyxl import load_workbook

    wb = load_workbook(written)
    assert wb.sheetnames[0] == "Extraction"
    assert "Observations" in wb.sheetnames
    assert "Metadata" in wb.sheetnames
    extract = wb["Extraction"]
    # Combined sheet must include observation headers + station values
    extract_values = [
        str(c) for row in extract.iter_rows(values_only=True) for c in row if c is not None
    ]
    extract_blob = " | ".join(extract_values)
    assert "Inst. Stn (from)" in extract_blob
    assert "SIAX-03" in extract_blob
    assert "363.054" in extract_blob
    assert "Victor Mokwenye" in extract_blob
    obs = wb["Observations"]
    headers = [c.value for c in obs[1]]
    assert "Inst. Stn (from)" in headers
    assert obs[2][0].value == "SIAX-03"
    assert "363.054" in str(obs[3][headers.index("Slope dist")].value)


def test_ocr_word_export_request_and_docx(tmp_path: Path) -> None:
    from agent.vision_ocr import (
        DOC_UI,
        export_ocr_extraction_to_docx,
        is_ocr_word_export_request,
        resolve_ocr_word_export_path,
        save_last_ocr_extraction,
    )

    assert is_ocr_word_export_request("save this into the file f.docx")
    assert is_ocr_word_export_request("save the extraction into the file f.docx")
    assert not is_ocr_word_export_request("save as a well-structured essay to f.docx")
    assert not is_ocr_word_export_request("save the updates to excel file e.xlsx")
    assert resolve_ocr_word_export_path("save this into the file f.docx", tmp_path).name == "f.docx"

    structured = {
        "document_type": DOC_UI,
        "title": "SurvyAI Desktop — 1.0.0",
        "metadata": {"workspace_path": "C:/Jobs/ODUOHA", "user": "Lasborne", "status": "Working"},
        "sections": [
            {"heading": "Conversations", "lines": ["Go to the only excel fil...", "New", "Delete"]},
            {"heading": "Live activity", "lines": ["[12:43:09] Task submitted.", "Elapsed: 2164s"]},
            {"heading": "Controls", "lines": ["Send", "Cancel", "Retry last"]},
        ],
    }
    saved = save_last_ocr_extraction(
        structured, workspace=tmp_path, image_paths=["Screenshot (247).png"], document_type=DOC_UI
    )
    assert saved is not None
    from agent.vision_ocr import load_last_ocr_extraction

    last = load_last_ocr_extraction(tmp_path)
    assert last is not None
    out = export_ocr_extraction_to_docx(
        last["structured"],
        tmp_path / "f.docx",
        image_paths=last.get("image_paths") or [],
        user_table=str(last.get("user_table") or ""),
    )
    assert out.is_file()
    import docx

    doc = docx.Document(str(out))
    texts = [p.text for p in doc.paragraphs if p.text.strip()]
    blob = "\n".join(texts)
    assert "SurvyAI Desktop" in blob
    assert "Conversations" in blob
    assert "Task submitted" in blob
    assert "Cancel" in blob
    # Must not be a meaningless filename-stem title only
    assert blob.strip() != "F"
    assert "Live activity" in blob


def test_ui_screenshot_classify_and_prompt() -> None:
    prompt = "Scan through the image attached and give all details"
    paths = ["Screenshot (247).png"]
    assert classify_ocr_document(prompt, paths) == DOC_UI
    assert is_ocr_only_request(prompt)
    user = _user_prompt_for_ocr(prompt, document_type=DOC_UI)
    assert "sections" in user.lower()
    assert "hz_fl" not in user
    system = _system_prompt_for_mode("structured", document_type=DOC_UI)
    assert "ui" in system.lower() or "sections" in system.lower()
    assert "Never return only" in system or "ALL readable" in system


def test_ui_screenshot_normalize_format_and_excel(tmp_path: Path) -> None:
    raw = {
        "document_type": "ui_screenshot",
        "title": "SurvyAI Desktop — 1.0.0",
        "metadata": {
            "workspace_path": "C:/Users/USER/Documents/JOB_301_replot/ODUOHA FAMILY",
            "user": "Lasborne",
            "status": "Working",
            "organization": None,
            "instrument": None,
        },
        "panels": [
            {
                "title": "Conversations",
                "items": [
                    "Go to the only excel fil...",
                    "Generate 'Check14.dw... Batch cadastral plott...",
                ],
            },
            {
                "heading": "Console / Chat",
                "lines": [
                    "YOU: Go to the only excel file contained in the folder...",
                    "Ask SurvyAI to create CAD drawings...",
                ],
            },
            {
                "name": "Live activity",
                "text": "[12:43:09] Task submitted.\n[12:43:09] Starting agent run...",
            },
            {
                "heading": "Controls",
                "lines": ["Send", "Cancel", "Retry last", "Use fallback LLM"],
            },
        ],
        "rows": [],
    }
    out = normalize_extracted_document(raw)
    assert out["document_type"] == DOC_UI
    assert len(out["sections"]) >= 4
    assert any(s["heading"] == "Conversations" for s in out["sections"])
    assert "organization" not in (out.get("metadata") or {}) or out["metadata"].get("organization") not in (None, "")

    repaired = apply_survey_context_repairs(out, document_type=DOC_UI)
    assert repaired.get("sections")
    assert not repaired.get("rows")

    user = VisionOcrResult(
        success=True, structured=out, image_paths=["Screenshot (247).png"], document_type=DOC_UI
    ).format_for_user()
    assert "SurvyAI Desktop" in user
    assert "### Conversations" in user
    assert "### Console / Chat" in user
    assert "### Live activity" in user
    assert "Task submitted" in user
    assert "Inst. Stn" not in user
    assert "Surveyed by" not in user
    assert "style_card" not in user
    # Must not be title-only
    assert user.count("###") >= 3

    review = build_ocr_review(
        image_paths=["Screenshot (247).png"],
        document_type=DOC_UI,
        model_name="test",
        structured=out,
    )
    assert review["sections"]
    assert review["sections"][0]["heading"]

    xlsx = export_ocr_extraction_to_excel(out, tmp_path / "ui.xlsx", image_paths=["Screenshot (247).png"])
    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    assert "Sections" in wb.sheetnames
    sec = wb["Sections"]
    blob = " | ".join(str(c) for row in sec.iter_rows(values_only=True) for c in row if c is not None)
    assert "Conversations" in blob
    assert "Live activity" in blob
    assert "Cancel" in blob
